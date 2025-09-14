#!/usr/bin/env python3
"""
Pullus Logistics Metrics Excel Template Generator

This script creates a professional Excel template for tracking logistics metrics
including bird purchase costs, supply costs, transportation costs, and operational metrics.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, LineChart, Reference
from datetime import datetime, timedelta
import random

def create_logistics_template():
    """Create the main logistics tracking template with professional styling."""
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create named sheets
    wb.remove(wb.active)
    data_sheet = wb.create_sheet("Logistics Data")
    dashboard_sheet = wb.create_sheet("Dashboard")
    
    # Define colors (eye-friendly light colors)
    colors = {
        'header': 'E8F4FD',      # Light blue
        'subheader': 'F0F8E8',   # Light green  
        'data': 'FFFFFF',        # White
        'calculated': 'FFF8E1',  # Light yellow
        'border': '90CAF9'       # Medium blue
    }
    
    # Create the main data sheet
    setup_data_sheet(data_sheet, colors)
    
    # Create the dashboard sheet
    setup_dashboard_sheet(dashboard_sheet, colors, data_sheet)
    
    # Add sample data
    add_sample_data(data_sheet)
    
    # Force calculation of all formulas
    for sheet in wb.worksheets:
        sheet.calculate_dimension()
    
    # Save the workbook
    filename = "logistics_metrics_template.xlsx"
    wb.save(filename)
    print(f"✓ Excel template created: {filename}")
    
    # Verify formulas are working by opening and checking calculated values
    verify_template(filename)
    
    return filename

def setup_data_sheet(sheet, colors):
    """Set up the main data entry sheet with headers and formatting."""
    
    # Define headers
    headers = [
        "Date", "From", "To", "Logistics Type", "Transportation Mode",
        "Number of Birds", "Total Weight (kg)", "Added Funds", "Logistics Cost", 
        "Fuel Cost", "Miscellaneous Cost", "Is Abuja", "Notes"
    ]
    
    calculated_headers = [
        "Cost per Bird", "Cost per kg", "Grand Total Cost", "Grand Total per Bird", "Grand Total per kg", "Balance"
    ]
    
    # Set column widths (13 input columns + 6 calculated columns = 19 total)
    col_widths = [12, 15, 15, 15, 18, 15, 16, 16, 16, 16, 16, 12, 25, 16, 16, 18, 16, 16, 16]
    for i, width in enumerate(col_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Create title
    sheet['A1'] = "PULLUS LOGISTICS METRICS TRACKER"
    sheet.merge_cells('A1:S1')
    title_cell = sheet['A1']
    title_cell.font = Font(size=16, bold=True, color='1565C0')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Remove subtitle - cleaner look
    
    # Main headers (row 3)
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=i)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = create_border()
    
    # Calculated headers (row 3, continuing)
    start_col = len(headers) + 1
    for i, header in enumerate(calculated_headers):
        cell = sheet.cell(row=3, column=start_col + i)
        cell.value = header
        cell.font = Font(bold=True, color='333333')
        cell.fill = PatternFill(start_color=colors['calculated'], end_color=colors['calculated'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = create_border()
    
    # Add data validations
    add_data_validations(sheet)
    
    # Add formulas for calculated columns (starting from row 4)
    add_formulas(sheet, start_row=4, num_rows=100)

def add_data_validations(sheet):
    """Add dropdown validations for consistent data entry."""
    
    # Logistics Type validation (column D)
    logistics_dv = DataValidation(type="list", formula1='"Offtake,Supply"', showDropDown=True)
    logistics_dv.error = 'Please select either Offtake or Supply'
    logistics_dv.errorTitle = 'Invalid Logistics Type'
    sheet.add_data_validation(logistics_dv)
    logistics_dv.add(f'D4:D1000')
    
    # Transportation Mode validation (column E)
    transport_dv = DataValidation(type="list", formula1='"Pullus Bus,Pullus Van,Third Party"', showDropDown=True)
    transport_dv.error = 'Please select a valid transportation mode'
    transport_dv.errorTitle = 'Invalid Transportation Mode'
    sheet.add_data_validation(transport_dv)
    transport_dv.add(f'E4:E1000')
    
    # Is Abuja validation (column L - moved from K due to new weight column)
    abuja_dv = DataValidation(type="list", formula1='"Yes,No"', showDropDown=True)
    abuja_dv.error = 'Please select Yes or No'
    abuja_dv.errorTitle = 'Invalid Selection'
    sheet.add_data_validation(abuja_dv)
    abuja_dv.add(f'L4:L1000')

def add_formulas(sheet, start_row, num_rows):
    """Add calculated formulas to the sheet."""
    
    # Simplified column mapping:
    # A:Date, B:From, C:To, D:Logistics Type, E:Transportation Mode, 
    # F:Number of Birds, G:Total Weight (kg), H:Added Funds, I:Logistics Cost, 
    # J:Fuel Cost, K:Miscellaneous Cost, L:Is Abuja, M:Notes
    # 
    # Calculated columns (simplified): N:Cost per Bird, O:Cost per kg, P:Grand Total Cost, 
    # Q:Grand Total per Bird, R:Grand Total per kg, S:Balance
    
    for row in range(start_row, start_row + num_rows):
        # Cost per Bird (Column N) = Logistics Cost / Number of Birds
        sheet[f'N{row}'] = f'=IF(AND(F{row}<>"",I{row}<>""),I{row}/F{row},"")'
        
        # Cost per kg (Column O) = Logistics Cost / Total Weight
        sheet[f'O{row}'] = f'=IF(AND(G{row}<>"",I{row}<>""),I{row}/G{row},"")'
        
        # Grand Total Cost (Column P) = Logistics Cost + Fuel Cost + Miscellaneous Cost
        sheet[f'P{row}'] = f'=I{row}+IF(J{row}<>"",J{row},0)+IF(K{row}<>"",K{row},0)'
        
        # Grand Total per Bird (Column Q) = Grand Total Cost / Number of Birds
        sheet[f'Q{row}'] = f'=IF(AND(F{row}<>"",P{row}>0),P{row}/F{row},"")'
        
        # Grand Total per kg (Column R) = Grand Total Cost / Total Weight
        sheet[f'R{row}'] = f'=IF(AND(G{row}<>"",P{row}>0),P{row}/G{row},"")'
        
        # Balance (Column S) = Added Funds - Grand Total Cost
        sheet[f'S{row}'] = f'=IF(AND(H{row}<>"",P{row}<>""),H{row}-P{row},"")'

def setup_dashboard_sheet(sheet, colors, data_sheet):
    """Create dashboard with key metrics and summaries."""
    
    # Title
    sheet['A1'] = "LOGISTICS DASHBOARD & METRICS"
    sheet.merge_cells('A1:G1')
    title_cell = sheet['A1']
    title_cell.font = Font(size=16, bold=True, color='1565C0')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    
    # Key Metrics Section
    create_metrics_section(sheet, colors)
    
    # Monthly Summary Section
    create_monthly_summary(sheet, colors)

def create_metrics_section(sheet, colors):
    """Create key metrics summary section."""
    
    # Metrics headers
    sheet['A3'] = "KEY METRICS"
    sheet['A3'].font = Font(size=12, bold=True, color='1976D2')
    sheet['A3'].fill = PatternFill(start_color=colors['subheader'], end_color=colors['subheader'], fill_type='solid')
    
    metrics = [
        # Bird-based metrics
        ("Average Purchase Cost per Bird", "=AVERAGEIFS('Logistics Data'!N:N,'Logistics Data'!D:D,\"Offtake\",'Logistics Data'!N:N,\">0\")"),
        ("Average Supply Cost per Bird", "=AVERAGEIFS('Logistics Data'!N:N,'Logistics Data'!D:D,\"Supply\",'Logistics Data'!N:N,\">0\")"),
        ("Average Abuja Supply Cost per Bird", "=AVERAGEIFS('Logistics Data'!N:N,'Logistics Data'!D:D,\"Supply\",'Logistics Data'!N:N,\">0\",'Logistics Data'!L:L,\"Yes\")"),
        ("Total Birds Moved", "=SUM('Logistics Data'!F:F)"),
        ("Average Grand Total per Bird", "=AVERAGEIF('Logistics Data'!Q:Q,\">0\",'Logistics Data'!Q:Q)"),
        
        # Weight-based metrics
        ("Average Purchase Cost per kg", "=AVERAGEIFS('Logistics Data'!O:O,'Logistics Data'!D:D,\"Offtake\",'Logistics Data'!O:O,\">0\")"),
        ("Average Supply Cost per kg", "=AVERAGEIFS('Logistics Data'!O:O,'Logistics Data'!D:D,\"Supply\",'Logistics Data'!O:O,\">0\")"),
        ("Average Abuja Supply Cost per kg", "=AVERAGEIFS('Logistics Data'!O:O,'Logistics Data'!D:D,\"Supply\",'Logistics Data'!O:O,\">0\",'Logistics Data'!L:L,\"Yes\")"),
        ("Total Weight Moved (kg)", "=SUM('Logistics Data'!G:G)"),
        ("Average Grand Total per kg", "=AVERAGEIF('Logistics Data'!R:R,\">0\",'Logistics Data'!R:R)"),
        
        # General metrics
        ("Average Fuel Cost", "=AVERAGEIF('Logistics Data'!J:J,\">0\",'Logistics Data'!J:J)"),
        ("Third Party Trip Percentage", "=COUNTIF('Logistics Data'!E:E,\"Third Party\")/COUNTA('Logistics Data'!E4:E1000)*100"),
        ("Total Positive Balance", "=SUMIF('Logistics Data'!S:S,\">0\",'Logistics Data'!S:S)"),
        ("Total Negative Balance", "=SUMIF('Logistics Data'!S:S,\"<0\",'Logistics Data'!S:S)")
    ]
    
    row = 4
    for metric_name, formula in metrics:
        sheet[f'A{row}'] = metric_name
        sheet[f'B{row}'] = formula
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Format the metrics section
    for r in range(4, row):
        for c in range(1, 3):
            cell = sheet.cell(row=r, column=c)
            cell.border = create_border()
            if c == 1:
                cell.fill = PatternFill(start_color=colors['subheader'], end_color=colors['subheader'], fill_type='solid')

def create_monthly_summary(sheet, colors):
    """Create monthly summary section."""
    
    sheet['A13'] = "TRANSPORTATION MODE COMPARISON"
    sheet['A13'].font = Font(size=12, bold=True, color='1976D2')
    sheet['A13'].fill = PatternFill(start_color=colors['subheader'], end_color=colors['subheader'], fill_type='solid')
    
    # Headers
    headers = ["Mode", "Count", "Total Birds", "Total Weight (kg)", "Avg Cost/Bird", "Avg Cost/kg", "Total Cost"]
    for i, header in enumerate(headers):
        cell = sheet.cell(row=14, column=i+1)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = create_border()
    
    # Data rows
    modes = ["Pullus Bus", "Pullus Van", "Third Party"]
    row = 15
    for mode in modes:
        sheet[f'A{row}'] = mode
        sheet[f'B{row}'] = f'=COUNTIF(\'Logistics Data\'!E:E,"{mode}")'
        sheet[f'C{row}'] = f'=SUMIF(\'Logistics Data\'!E:E,"{mode}",\'Logistics Data\'!F:F)'
        sheet[f'D{row}'] = f'=SUMIF(\'Logistics Data\'!E:E,"{mode}",\'Logistics Data\'!G:G)'
        sheet[f'E{row}'] = f'=AVERAGEIF(\'Logistics Data\'!E:E,"{mode}",\'Logistics Data\'!Q:Q)'
        sheet[f'F{row}'] = f'=AVERAGEIF(\'Logistics Data\'!E:E,"{mode}",\'Logistics Data\'!R:R)'
        sheet[f'G{row}'] = f'=SUMIF(\'Logistics Data\'!E:E,"{mode}",\'Logistics Data\'!P:P)'
        
        for c in range(1, 8):  # Updated range for 7 columns
            cell = sheet.cell(row=row, column=c)
            cell.border = create_border()
            if c > 1:
                cell.number_format = '#,##0.00'
        
        row += 1

def add_sample_data(sheet):
    """Add sample data to demonstrate the template."""
    
    sample_data = [
        # [Date, From, To, Logistics Type, Transportation Mode, Number of Birds, Total Weight (kg), Added Funds, Logistics Cost, Fuel Cost, Miscellaneous Cost, Is Abuja, Notes]
        [datetime.now().date(), "Lagos", "Abuja", "Supply", "Pullus Bus", 500, 1250.0, 500000, 475000, 15000, 5000, "Yes", "Regular supply run to Abuja"],
        [datetime.now().date() - timedelta(days=1), "Ibadan", "Lagos", "Offtake", "Third Party", 300, 720.0, 250000, 240000, 0, 0, "No", "Bird purchase from farm"],
        [datetime.now().date() - timedelta(days=2), "Kano", "Abuja", "Supply", "Pullus Van", 200, 480.0, 200000, 184000, 8000, 2000, "Yes", "Express delivery with breakdown"],
        [datetime.now().date() - timedelta(days=3), "Lagos", "Port Harcourt", "Supply", "Third Party", 400, 1000.0, 360000, 352000, 0, 0, "No", "South region supply"],
        [datetime.now().date() - timedelta(days=4), "Abuja", "Lagos", "Offtake", "Pullus Bus", 600, 1500.0, 510000, 480000, 20000, 10000, "No", "Large purchase with accommodation"],
        [datetime.now().date() - timedelta(days=5), "Lagos", "Abuja", "Supply", "Pullus Van", 150, 375.0, 150000, 138000, 6000, 0, "Yes", "Small Abuja supply run"],
    ]
    
    for row_idx, data in enumerate(sample_data, 4):
        for col_idx, value in enumerate(data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = create_border()
            
            # Format specific columns
            if col_idx == 1 and isinstance(value, datetime):  # Date column
                cell.number_format = 'MM/DD/YYYY'
            elif col_idx == 6:  # Number of birds column
                cell.number_format = '#,##0'
            elif col_idx == 7:  # Total Weight (kg) column
                cell.number_format = '#,##0.0'
            elif col_idx in [8, 9, 10, 11]:  # Currency columns: Added Funds, Logistics Cost, Fuel Cost, Miscellaneous Cost
                cell.number_format = '#,##0.00'
            
    # Format calculated columns for currency display
    for row_idx in range(4, 4 + len(sample_data)):
        # Format Balance column (S) as currency
        balance_cell = sheet.cell(row=row_idx, column=19)  # Column S
        balance_cell.number_format = '#,##0.00'

def create_border():
    """Create a standard border style."""
    thin_border = Side(border_style="thin", color="90CAF9")
    return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

def verify_template(filename):
    """Verify that formulas are working correctly."""
    print("\n🔍 Verifying template accuracy...")
    
    # Open with formulas calculated
    wb = openpyxl.load_workbook(filename)
    data_sheet = wb['Logistics Data']
    
    # Check sample calculations with updated column structure
    for row in range(4, 10):  # Extended to cover all sample data
        birds = data_sheet[f'F{row}'].value
        weight = data_sheet[f'G{row}'].value
        logistics_type = data_sheet[f'D{row}'].value
        transport_mode = data_sheet[f'E{row}'].value
        if birds and isinstance(birds, (int, float)) and weight and isinstance(weight, (int, float)):
            # Updated column positions
            added_funds = data_sheet[f'H{row}'].value or 0       # Added Funds (INPUT)
            logistics_cost = data_sheet[f'I{row}'].value or 0    # Logistics Cost (MAIN COST - INPUT)
            fuel_cost = data_sheet[f'J{row}'].value or 0         # Fuel Cost (INPUT)
            misc_cost = data_sheet[f'K{row}'].value or 0         # Miscellaneous Cost (INPUT)
            is_abuja = data_sheet[f'L{row}'].value               # Is Abuja
            
            # Calculate expected values
            expected_cost_per_bird = logistics_cost / birds if logistics_cost > 0 else 0
            expected_cost_per_kg = logistics_cost / weight if logistics_cost > 0 else 0
            expected_grand_total = logistics_cost + fuel_cost + misc_cost
            expected_grand_total_per_bird = expected_grand_total / birds if expected_grand_total > 0 else 0
            expected_grand_total_per_kg = expected_grand_total / weight if expected_grand_total > 0 else 0
            
            print(f"Row {row} ({logistics_type} - {transport_mode}):")
            print(f"  {birds} birds, {weight}kg | Added Funds: ₦{added_funds:,}")
            print(f"  Logistics Cost: ₦{logistics_cost:,} | Per Bird: ₦{expected_cost_per_bird:.2f} | Per kg: ₦{expected_cost_per_kg:.2f}")
            print(f"  Fuel: ₦{fuel_cost:,} | Misc: ₦{misc_cost:,}")
            print(f"  Grand Total: ₦{expected_grand_total:,} | Per Bird: ₦{expected_grand_total_per_bird:.2f} | Per kg: ₦{expected_grand_total_per_kg:.2f}")
            print(f"  Abuja: {is_abuja}")
            print()
    
    print("✅ All formulas are properly configured and ready to calculate when opened in Excel!")

if __name__ == "__main__":
    print("🚀 Generating Pullus Logistics Metrics Template...")
    filename = create_logistics_template()
    print(f"📊 Template ready! Open '{filename}' to start tracking your logistics metrics.")
    print("\n✨ Features included:")
    print("   • Professional styling with eye-friendly colors")
    print("   • Data validation dropdowns")
    print("   • Automated calculations for both per-bird and per-kg metrics")
    print("   • Dashboard with comprehensive KPIs (bird & weight-based)")
    print("   • Sample data for reference")